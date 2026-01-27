import os
import sys
import json
import logging
import uuid
from pathlib import Path
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl.utils import range_boundaries
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage

# --- 配置区域 ---
LM_STUDIO_API_BASE = "http://192.168.30.190:1234/v1"
LM_STUDIO_API_KEY = "lm-studio"
MODEL_NAME = "qwen/qwen3-4b-2507"
TEMPERATURE = 0.1
MAX_TOKENS = 32000
OUTPUT_DIR = r"D:\DM\LangChain\联络名单-Agent\json_save"

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('excel_agent.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

class ExcelProcessingAgent:
    """
    基于LLM的Excel智能处理Agent
    负责：表头识别、合并单元格处理、数据分块、JSON转换与持久化
    """
    
    def __init__(self):
        self.llm = self._get_llm_client()
        self.output_dir = Path(OUTPUT_DIR)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Agent 初始化完成，输出目录: {self.output_dir}")

    def _get_llm_client(self) -> ChatOpenAI:
        """初始化 LangChain 客户端"""
        return ChatOpenAI(
            model=MODEL_NAME,
            openai_api_base=LM_STUDIO_API_BASE,
            openai_api_key=LM_STUDIO_API_KEY,
            temperature=TEMPERATURE,
            max_tokens=None if MAX_TOKENS == -1 else MAX_TOKENS
        )

    def identify_header_row(self, sheet, sheet_name: str) -> int:
        """
        【表头识别Agent】
        截取前20行，询问LLM哪一行是表头
        """
        logger.info(f"正在识别工作表 '{sheet_name}' 的表头...")
        
        # 1. 获取前20行数据预览
        preview_rows = []
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True)):
            if any(cell is not None for cell in row):
                preview_rows.append(f"行号 {i}: {list(row)}")
        
        preview_text = "\n".join(preview_rows)
        
        # 2. 构造提示词
        prompt = f"""
你是一个专业的Excel数据分析助手。请分析以下Excel工作表的前20行数据，找出包含列名（表头）的那一行。
通常表头行包含描述性的字段名称（如"姓名"、"电话"、"地址"等）。

数据预览：
{preview_text}

请直接返回表头所在的行号（从0开始计数）。
只返回一个数字，不要包含任何其他文字或解释。
如果无法确定，请返回 0。
"""
        try:
            response = self.llm.invoke([HumanMessage(content=prompt)])
            content = response.content.strip()
            import re
            match = re.search(r'\d+', content)
            if match:
                header_index = int(match.group())
                logger.info(f"LLM 识别表头在第 {header_index} 行")
                return header_index
            else:
                logger.warning("LLM 未返回有效数字，默认使用第 0 行")
                return 0
        except Exception as e:
            logger.error(f"表头识别失败: {e}，默认使用第 0 行")
            return 0

    def handle_merged_cells(self, sheet, header_row_idx: int):
        """
        处理合并单元格：拆分并填充值
        """
        logger.info("正在处理合并单元格...")
        merged_ranges = list(sheet.merged_cells.ranges)
        
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            
            sheet.unmerge_cells(str(merged_range))
            
            is_header_area = (min_row - 1) == header_row_idx
            
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.value = top_left_cell_value

    def unique_headers(self, headers: List[str]) -> List[str]:
        """对表头进行去重处理，空值转为Unknown"""
        seen = {}
        new_headers = []
        for h in headers:
            h_str = str(h).strip() if h is not None else "Unknown"
            if not h_str: # Handle empty string
                h_str = "Unknown"
                
            if h_str in seen:
                seen[h_str] += 1
                new_headers.append(f"{h_str}_{seen[h_str]}")
            else:
                seen[h_str] = 0
                new_headers.append(h_str)
        return new_headers

    def process_chunk_with_llm(self, headers: List[str], data_rows: List[List[Any]]) -> List[Dict]:
        """
        【智能分块Agent】
        将数据块转换为标准JSON
        """
        # 构造数据预览
        data_preview = []
        for row in data_rows:
            # 确保行长度与表头一致 (调用前已过滤列，这里长度应该匹配)
            row_data = [str(val) if val is not None else "" for val in row]
            # 截断过长的数据
            row_data = [s[:50] + "..." if len(s) > 50 else s for s in row_data]
            data_preview.append(str(row_data))
            
        data_str = "\n".join(data_preview)
        
        prompt = f"""
你是一个数据转换专家。请将以下Excel数据行转换为标准的JSON格式。
表头列表：{headers}

数据行（每行对应一个JSON对象）：
{data_str}

要求：
1. 返回一个JSON数组。
2. 每个对象包含所有表头字段。
3. 保持原始数据类型（数字转为数字，文本转为文本）。
4. 只返回JSON代码，不要包含Markdown格式。
"""
        try:
            response = self.llm.invoke([HumanMessage(content=prompt)])
            content = response.content.strip()
            if content.startswith("```json"):
                content = content[7:]
            if content.startswith("```"):
                content = content[3:]
            if content.endswith("```"):
                content = content[:-3]
            
            return json.loads(content)
        except json.JSONDecodeError:
            logger.error("LLM 返回的不是有效JSON，尝试手动降级处理")
            result = []
            for row in data_rows:
                item = {}
                for i, h in enumerate(headers):
                    val = row[i] if i < len(row) else None
                    item[h] = val
                result.append(item)
            return result
        except Exception as e:
            logger.error(f"分块处理异常: {e}")
            return []

    def process_sheet(self, file_path: Path, wb, sheet_name: str):
        """处理单个工作表"""
        sheet = wb[sheet_name]
        
        # 1. 识别表头
        header_row_idx = self.identify_header_row(sheet, sheet_name)
        
        # 2. 处理合并单元格
        self.handle_merged_cells(sheet, header_row_idx)
        
        # 3. 获取并清洗表头
        header_row_values = []
        for cell in sheet[header_row_idx + 1]:
            header_row_values.append(cell.value)
        
        full_headers = self.unique_headers(header_row_values)
        logger.info(f"原始表头: {full_headers}")
        
        # 清洗：找出非 'Unknown' 的列索引
        valid_indices = []
        cleaned_headers = []
        for idx, h in enumerate(full_headers):
            if not h.startswith('Unknown'):
                valid_indices.append(idx)
                cleaned_headers.append(h)
                
        logger.info(f"清洗后的表头 (保留 {len(cleaned_headers)} 列): {cleaned_headers}")
        
        # 4. 准备输出文件
        safe_filename = file_path.stem
        safe_sheet_name = "".join([c if c.isalnum() or c in (' ', '_', '-') else '_' for c in sheet_name])
        output_file = self.output_dir / f"{safe_filename}_{safe_sheet_name}.json"
        
        # 生成文档ID (紧凑型UUID)
        doc_id = f"doc-{uuid.uuid4().hex}"
        
        # 断点续传检查
        processed_count = 0
        if output_file.exists():
            try:
                with open(output_file, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
                    processed_count = len(existing_data)
                    logger.info(f"发现已有进度，已处理 {processed_count} 个分块，将跳过。")
            except json.JSONDecodeError:
                logger.warning("已有文件损坏，将重新开始。")
                processed_count = 0

        # 5. 分块处理数据
        data_start_row = header_row_idx + 2 
        current_batch = []
        
        # 初始化文件
        if processed_count == 0:
             with open(output_file, 'w', encoding='utf-8') as f:
                 f.write("[\n")
        else:
            # 修正文件结尾，准备追加
            with open(output_file, 'rb+') as f:
                f.seek(-1, os.SEEK_END)
                while f.read(1) != b']':
                    f.seek(-2, os.SEEK_CUR)
                f.seek(-1, os.SEEK_CUR)
                f.truncate()
        
        # 统计已保存的CHUNK数量
        chunks_saved = 0
        
        row_iterator = sheet.iter_rows(min_row=data_start_row, values_only=True)
        
        # 批处理缓存
        rows_buffer = []
        
        for row in row_iterator:
            # 跳过全空行
            if all(cell is None for cell in row):
                continue
            
            # 提取有效列数据
            filtered_row = [row[i] for i in valid_indices]
            rows_buffer.append(filtered_row)
            
            if len(rows_buffer) >= 10:
                # 检查是否需要跳过（断点续传）
                # 注意：processed_count 是已保存的 chunk 数量
                if chunks_saved < processed_count:
                    chunks_saved += 1
                    rows_buffer = [] # 清空buffer，模拟已处理
                    continue
                
                # 处理当前块
                self._process_and_save_chunk(cleaned_headers, rows_buffer, doc_id, output_file, is_first=(chunks_saved==0 and processed_count==0))
                chunks_saved += 1
                rows_buffer = []
                logger.info(f"工作表 {sheet_name}: 已保存 {chunks_saved} 个分块")

        # 处理剩余数据
        if rows_buffer:
             if chunks_saved >= processed_count:
                self._process_and_save_chunk(cleaned_headers, rows_buffer, doc_id, output_file, is_first=(chunks_saved==0 and processed_count==0))
                chunks_saved += 1

        # 封闭JSON数组
        with open(output_file, 'a', encoding='utf-8') as f:
            f.write("\n]")
            
        logger.info(f"工作表 {sheet_name} 处理完成，共保存 {chunks_saved} 个分块。")

    def _process_and_save_chunk(self, headers, rows, doc_id, output_file, is_first):
        """生成分块结构并写入"""
        # 1. LLM 转换内容
        content_list = self.process_chunk_with_llm(headers, rows)
        
        # 2. 生成 Chunk ID
        chunk_id = f"chunk-{uuid.uuid4().hex}"
        
        # 3. 构造新结构
        wrapper = {
            chunk_id: {
                "content": content_list,
                "doc_id": doc_id,
                "chunk_id": chunk_id
            }
        }
        
        # 4. 写入
        self._append_json_to_file(output_file, [wrapper], is_first_batch=is_first)

    def _append_json_to_file(self, file_path, data_list, is_first_batch=False):
        """原子追加写入JSON"""
        if not data_list:
            return
            
        json_strings = []
        for item in data_list:
            json_strings.append(json.dumps(item, ensure_ascii=False, indent=2))
            
        content = ",\n".join(json_strings)
        
        with open(file_path, 'a', encoding='utf-8') as f:
            if not is_first_batch:
                f.write(",\n")
            f.write(content)

    def process_file(self, file_path_str: str):
        """主入口：处理单个Excel文件"""
        file_path = Path(file_path_str)
        if not file_path.exists():
            logger.error(f"文件不存在: {file_path}")
            return

        logger.info(f"开始处理文件: {file_path}")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                try:
                    self.process_sheet(file_path, wb, sheet_name)
                except Exception as e:
                    logger.error(f"处理工作表 {sheet_name} 失败: {e}", exc_info=True)
            wb.close()
            logger.info(f"文件 {file_path} 处理完毕")
        except Exception as e:
            logger.error(f"文件加载失败: {e}", exc_info=True)

if __name__ == "__main__":
    agent = ExcelProcessingAgent()
    print("=== Excel 智能分块处理 Agent (RAG Optimized) ===")
    target_input = input("请输入Excel文件路径或文件夹路径: ").strip()
    
    if not target_input:
        print("未输入路径，程序退出。")
        sys.exit(0)
        
    target_path = Path(target_input)
    if target_path.is_file():
        agent.process_file(str(target_path))
    elif target_path.is_dir():
        excel_files = list(target_path.rglob("*.xlsx"))
        print(f"找到 {len(excel_files)} 个Excel文件。")
        for f in excel_files:
            if not f.name.startswith("~$"):
                agent.process_file(str(f))
    else:
        print("无效的路径。")
